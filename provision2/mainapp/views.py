from django.views import View
from django.views.generic import ListView
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.core.serializers.json import DjangoJSONEncoder
from django.urls import reverse
from django.db.models import Q, ProtectedError
from django.core.paginator import Paginator
from .models import Fornitore, Magazzino, Mezzo, Tipologia, Zona, Listino, InserimentoFallito
from openpyxl import load_workbook
import pandas as pd
import json
import io
from django.shortcuts import render
from django.contrib import messages
from .tools import *
from django.contrib.auth.decorators import user_passes_test
from datetime import date
from django.http import JsonResponse


# -------------------------------------------------------------------------------------------------------------- #
# ---------------------------------- ANAGRAFICA FORNITORI ------------------------------------------------------ #
# -------------------------------------------------------------------------------------------------------------- #

# View per la visualizzazione e l'aggiunta di un fornitore 
def situazione_fornitori(request):
    # Ottieni tutti i fornitori ordinati per nome
    all_fornitori = Fornitore.objects.all().order_by('fornitore_nome')
    
    # Imposta la paginazione
    paginator = Paginator(all_fornitori, 15)  # 15 fornitori per pagina
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj
    }

    if request.method == 'POST':
        # Gestione dell'aggiunta di un nuovo fornitore
        nome = request.POST.get('fornitore_nome').strip().upper()
        cod_as = request.POST.get('fornitore_cod_as').strip().upper()

        if not nome or not cod_as:
            messages.error(request, 'I campi non possono essere vuoti')
        elif Fornitore.objects.filter(fornitore_cod_as=cod_as).exists():
            messages.error(request, 'Il codice AS è già presente')
        else:
            new_fornitore = Fornitore(fornitore_nome=nome, fornitore_cod_as=cod_as)
            new_fornitore.save()
            messages.success(request, 'Fornitore aggiunto con successo')
            return redirect('mainapp:situazione_fornitori')

    return render(request, 'mainapp/situazione_fornitori.html', context)

# View per l'eliminazione di un fornitore 
def elimina_fornitore(request, pk):
    if request.method == 'POST':
        fornitore = get_object_or_404(Fornitore, pk=pk)
        try:
            fornitore.delete()
            messages.success(request, 'Fornitore eliminato con successo')
        except ProtectedError:
            listini_associati = Listino.objects.filter(fornitore=fornitore).count()
            messages.error(request, f'Impossibile eliminare il fornitore. Ci sono {listini_associati} listini associati.')
    
    return redirect('mainapp:situazione_fornitori')


# View per la modifica di un fornitore
# CONTROLLO SULL'UNICITà DEL NOME DISABILITATO
def modifica_fornitore(request, pk):
    fornitore = get_object_or_404(Fornitore, pk=pk)
    if request.method == 'POST':
        nuovo_nome = request.POST.get('fornitore_nome', '').strip().upper()
        nuovo_codice = request.POST.get('fornitore_cod_as', '').strip().upper()

        if nuovo_nome and nuovo_codice:
            if not check_esistenza_fornitore_as(nuovo_codice):
                fornitore.fornitore_nome = nuovo_nome
                fornitore.fornitore_cod_as = nuovo_codice
                fornitore.save()
                messages.success(request, 'Fornitore modificato con successo')
            else:
                messages.error(request, 'Errore: Il codice fornitore ')
        
        elif nuovo_nome:
            fornitore.fornitore_nome = nuovo_nome
            fornitore.save()
            messages.success(request, 'Nome fornitore modificato con successo')
        elif nuovo_codice:
            if not check_esistenza_fornitore_as(nuovo_codice):
                fornitore.fornitore_cod_as = nuovo_codice
                fornitore.save()
                messages.success(request, 'Codice fornitore modificato con successo')
            else:
                messages.error(request, 'Errore: Codice fornitore già presente')  
        else:
            messages.error(request, 'Errore: Nessun dato inserito per la modifica')

    return redirect('mainapp:situazione_fornitori')
# -------------------------------------------------------------------------------------------------------------- #
# -------------------------------- FINE ANAGRAFICA FORNITORI --------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #



# -------------------------------------------------------------------------------------------------------------- #
# ----------------------------------- ANAGRAFICA SOCIETA' ------------------------------------------------------ #
# -------------------------------------------------------------------------------------------------------------- #
def situazione_societa(request):
    all_societa = Societa.objects.all().order_by('societa_nome')

    paginator = Paginator(all_societa, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj
    }
    if request.method == 'POST':
        # Gestione dell'aggiunta di una nuova società 
        nome = request.POST.get('societa_nome').strip().upper()

        if not nome:
            messages.error(request, 'Errore: inserire il nome della niova società')
        elif Societa.objects.filter(societa_nome=nome).exists():
            messages.error(request, 'Errore: La società esiste già')
        else:
            new_societa = Societa(societa_nome=nome)
            new_societa.save()
            messages.success(request, 'Società aggiunta con successo')
            return redirect('mainapp:situazione_societa')

    return render(request, 'mainapp/situazione_societa.html', context)

def elimina_societa(request, pk):
    if request.method == 'POST':
        societa = get_object_or_404(Societa, pk=pk)
        try:
            societa.delete()
            messages.success(request, 'Società eliminata con successo')
        except ProtectedError:
            messages.error(request, f'Impossibile eliminare la società di ID {societa.id }. Ci sono altre entità associate.')
    
    return redirect('mainapp:situazione_societa')

def modifica_societa(request, pk):
    societa = get_object_or_404(Societa, pk=pk)
    print(societa.pk)
    if request.method == 'POST':
        nuovo_nome = request.POST.get('societa_nome', '').strip().upper()
        print(nuovo_nome)
        if nuovo_nome:
            if not check_esistenza_societa(nuovo_nome):
                print("sono qua")
                societa.societa_nome = nuovo_nome
                societa.save()
                messages.success(request, 'Società modificata con successo')
            else:
                messages.error(request, 'Errore: Società già presente')
        else:
            messages.error(request, 'Errore: Nessun dato inserito per la modifica')

    return redirect('mainapp:situazione_societa')
# -------------------------------------------------------------------------------------------------------------- #
# --------------------------------- FINE ANAGRAFICA SOCIETA' --------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #



# -------------------------------------------------------------------------------------------------------------- #
# ---------------------------------- ANAGRAFICA MAGAZZINI ------------------------------------------------------ #
# -------------------------------------------------------------------------------------------------------------- #
def situazione_magazzini(request):
    all_magazzini = Magazzino.objects.all().order_by('magazzino_nome')
    societa_list = Societa.objects.all().order_by('societa_nome')

    paginator = Paginator(all_magazzini, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'societa_list': societa_list
    }

    if request.method == 'POST':
        nome = request.POST.get('magazzino_nome', '').strip().upper()
        lettera = request.POST.get('magazzino_lettera', '').strip().upper()
        societa_ids = request.POST.getlist('societa')  # Usa getlist invece di get

        if not nome or not lettera:
            messages.error(request, 'Errore: tutti i campi sono obbligatori')
        elif len(lettera) != 1:
            messages.error(request, 'Errore: La lettera del magazzino deve essere di un solo carattere')
        elif Magazzino.objects.filter(magazzino_lettera=lettera).exists():
            messages.error(request, 'Errore: La lettera del magazzino esiste già')
        else:
            try:
                new_magazzino = Magazzino(magazzino_nome=nome, magazzino_lettera=lettera)
                new_magazzino.save()
                for societa_id in societa_ids:
                    societa = Societa.objects.get(pk=societa_id)
                    new_magazzino.societa.add(societa)
                messages.success(request, 'Magazzino aggiunto con successo')
                return redirect('mainapp:situazione_magazzini')
            except Societa.DoesNotExist:
                messages.error(request, 'Errore: Società non valida')

    return render(request, 'mainapp/situazione_magazzini.html', context)

def modifica_magazzino(request, pk):
    magazzino = get_object_or_404(Magazzino, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('magazzino_nome', '').strip().upper()
        lettera = request.POST.get('magazzino_lettera', '').strip().upper()
        societa_id = request.POST.get('societa')

        if not nome or not lettera or not societa_id:
            messages.error(request, 'Errore: tutti i campi sono obbligatori')
        elif societa_id == "":
            messages.error(request, 'Errore: devi selezionare una società')
        elif len(lettera) != 1:
            messages.error(request, 'Errore: Il campo lettera deve contenere esattamente un carattere')
        elif Magazzino.objects.filter(magazzino_lettera=lettera).exclude(pk=pk).exists():
            messages.error(request, 'Errore: La lettera del magazzino esiste già')
        else:
            try:
                societa = Societa.objects.get(pk=societa_id)
                magazzino.magazzino_nome = nome
                magazzino.magazzino_lettera = lettera
                magazzino.societa = societa
                magazzino.save()
                messages.success(request, 'Magazzino modificato con successo')
            except Societa.DoesNotExist:
                messages.error(request, 'Errore: Società non valida')

    return redirect('mainapp:situazione_magazzini')

def elimina_magazzino(request, pk):
    if request.method == 'POST':
        magazzino = get_object_or_404(Magazzino, pk=pk)
        try:
            magazzino.delete()
            messages.success(request, 'Magazzino eliminato con successo')
        except Exception as e:
            messages.error(request, f'Errore durante l\'eliminazione del magazzino: {str(e)}')
    return redirect('mainapp:situazione_magazzini')
# -------------------------------------------------------------------------------------------------------------- #
# --------------------------------- FINE ANAGRAFICA MAGAZZINI -------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #



# -------------------------------------------------------------------------------------------------------------- #
# --------------------------------------- ANAGRAFICA MEZZI ----------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #
def situazione_mezzi(request):
    all_mezzi = Mezzo.objects.all().order_by('mezzo_nome')

    paginator = Paginator(all_mezzi, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }

    if request.method == 'POST':
        nome = request.POST.get('mezzo_nome', '').strip().upper()

        if not nome:
            messages.error(request, 'Errore: il nome del mezzo è obbligatorio')
        elif Mezzo.objects.filter(mezzo_nome=nome).exists():
            messages.error(request, 'Errore: Il mezzo esiste già')
        else:
            new_mezzo = Mezzo(mezzo_nome=nome)
            new_mezzo.save()
            messages.success(request, 'Mezzo aggiunto con successo')
            return redirect('mainapp:situazione_mezzi')

    return render(request, 'mainapp/situazione_mezzi.html', context)

def modifica_mezzo(request, pk):
    mezzo = get_object_or_404(Mezzo, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('mezzo_nome', '').strip().upper()

        if not nome:
            messages.error(request, 'Errore: il nome del mezzo è obbligatorio')
        elif Mezzo.objects.filter(mezzo_nome=nome).exclude(pk=pk).exists():
            messages.error(request, 'Errore: Il mezzo esiste già')
        else:
            mezzo.mezzo_nome = nome
            mezzo.save()
            messages.success(request, 'Mezzo modificato con successo')

    return redirect('mainapp:situazione_mezzi')

def elimina_mezzo(request, pk):
    if request.method == 'POST':
        mezzo = get_object_or_404(Mezzo, pk=pk)
        try:
            mezzo.delete()
            messages.success(request, 'Mezzo eliminato con successo')
        except Exception as e:
            messages.error(request, f'Errore durante l\'eliminazione del mezzo: {str(e)}')
    return redirect('mainapp:situazione_mezzi')
# -------------------------------------------------------------------------------------------------------------- #
# ------------------------------------- FINE ANAGRAFICA MEZZI -------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #


# -------------------------------------------------------------------------------------------------------------- #
# --------------------------------------- ANAGRAFICA TIPOLOGIE ------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #
def situazione_tipologie(request):
    all_tipologie = Tipologia.objects.all().order_by('tipologia_nome')

    paginator = Paginator(all_tipologie, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }

    if request.method == 'POST':
        nome = request.POST.get('tipologia_nome', '').strip().upper()

        if not nome:
            messages.error(request, 'Errore: il nome della tipologia è obbligatorio')
        elif Tipologia.objects.filter(tipologia_nome=nome).exists():
            messages.error(request, 'Errore: La tipologia esiste già')
        else:
            new_tipologia = Tipologia(tipologia_nome=nome)
            new_tipologia.save()
            messages.success(request, 'Tipologia aggiunta con successo')
            return redirect('mainapp:situazione_tipologie')

    return render(request, 'mainapp/situazione_tipologie.html', context)

def modifica_tipologia(request, pk):
    tipologia = get_object_or_404(Tipologia, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('tipologia_nome', '').strip().upper()

        if not nome:
            messages.error(request, 'Errore: il nome della tipologia è obbligatorio')
        elif Tipologia.objects.filter(tipologia_nome=nome).exclude(pk=pk).exists():
            messages.error(request, 'Errore: La tipologia esiste già')
        else:
            tipologia.tipologia_nome = nome
            tipologia.save()
            messages.success(request, 'Tipologia modificata con successo')

    return redirect('mainapp:situazione_tipologie')

def elimina_tipologia(request, pk):
    if request.method == 'POST':
        tipologia = get_object_or_404(Tipologia, pk=pk)
        try:
            tipologia.delete()
            messages.success(request, 'Tipologia eliminata con successo')
        except Exception as e:
            messages.error(request, f'Errore durante l\'eliminazione della tipologia: {str(e)}')
    return redirect('mainapp:situazione_tipologie')



# -------------------------------------------------------------------------------------------------------------- #
# --------------------------------- FINE ANAGRAFICA TIPOLOGIE -------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #


# -------------------------------------------------------------------------------------------------------------- #
# --------------------------------------- ANAGRAFICA ZONE ------------------------------------------------------ #
# -------------------------------------------------------------------------------------------------------------- #
def situazione_zone(request):
    all_zone = Zona.objects.all().order_by('zona_nome')

    paginator = Paginator(all_zone, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }

    if request.method == 'POST':
        nome = request.POST.get('zona_nome', '').strip().upper()

        if not nome:
            messages.error(request, 'Errore: il nome della zona è obbligatorio')
        elif Zona.objects.filter(zona_nome=nome).exists():
            messages.error(request, 'Errore: La zona esiste già')
        else:
            new_zona = Zona(zona_nome=nome)
            new_zona.save()
            messages.success(request, 'Zona aggiunta con successo')
            return redirect('mainapp:situazione_zone')

    return render(request, 'mainapp/situazione_zone.html', context)

def modifica_zona(request, pk):
    zona = get_object_or_404(Zona, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('zona_nome', '').strip().upper()

        if not nome:
            messages.error(request, 'Errore: il nome della zona è obbligatorio')
        elif Zona.objects.filter(zona_nome=nome).exclude(pk=pk).exists():
            messages.error(request, 'Errore: La zona esiste già')
        else:
            zona.zona_nome = nome
            zona.save()
            messages.success(request, 'Zona modificata con successo')

    return redirect('mainapp:situazione_zone')

def elimina_zona(request, pk):
    if request.method == 'POST':
        zona = get_object_or_404(Zona, pk=pk)
        try:
            zona.delete()
            messages.success(request, 'Zona eliminata con successo')
        except Exception as e:
            messages.error(request, f'Errore durante l\'eliminazione della zona: {str(e)}')
    return redirect('mainapp:situazione_zone')


class ImportZone(View):
    template_name = "mainapp/import_entita.html"

    def get(self, request):
        context = {'entita': 'zone'}
        return render(request, self.template_name, context=context)
    def post(self, request):
        if 'file' not in request.FILES:
            messages.error(request, 'Nessun file caricato')
            return redirect('mainapp:import_zone')
        
        excel_file = request.FILES['file']
        print(excel_file)
        
        # Leggo il file in memoria
        try:
            file_in_memory = io.BytesIO(excel_file.read())
            
            # carico il workbook
            wb = load_workbook(filename=file_in_memory, read_only=True, data_only=True)
            
            # primo foglio
            sheet = wb.active

            # converto il foglio in un DataFrame pandas
            data = sheet.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)
        except Exception as e:
            print("Errore nel leggere il file")
            print(f"Errore specifico: {type(e).__name__}, {str(e)}")
            messages.error(request, f'Errore nel leggere il file Excel: {str(e)}')
            return redirect('mainapp:import_zone')

        # Verifica che il file abbia tutte le colonne
        required_columns = ['ZONA' ]
        if not all(col in df.columns for col in required_columns):
            messages.error(request, 'Il file non contiene tutte le colonne richieste')
            return redirect('mainapp:import_zone')

        righe_inserite = 0
        righe_fallite = 0

        for _,row in df.iterrows():
            
            with transaction.atomic():
                new_zona_nome = row['ZONA'].strip().upper()
                
                if check_esistenza_zona(new_zona_nome):
                    righe_fallite = righe_fallite + 1
                    print(f'La zona {new_zona_nome} esiste già ')
                else:
                    zona = Zona.objects.create(zona_nome=new_zona_nome)
                    righe_inserite = righe_inserite + 1

        if righe_fallite > 0:
            messages.warning(request, f'Importazione completata. Inserite: {righe_inserite}, Fallite: {righe_fallite}')
        else:
            messages.success(request, f'Importazione completata. Inserite: {righe_inserite}, Fallite: {righe_fallite}')

        return redirect('mainapp:import_zone')             
             
# -------------------------------------------------------------------------------------------------------------- #
# -------------------------------------- FINE ANAGRAFICA ZONE -------------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------- #


# -------------------------------------------------------------------------------------------------------------- #
# ---------------------------- GESTIONE CARICAMENTO DEL LISTINO ATTRAVERSO FILE EXCEL -------------------------- #
# -------------------------------------------------------------------------------------------------------------- #
class CustomJSONEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, pd.Timestamp):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        return super().default(obj)
    


class ImportListinoView(View):
    template_name = 'mainapp/import_listino.html'

    def get(self, request):
        return render(request, self.template_name)
    
    def post(self, request):
        if 'file' not in request.FILES:
            messages.error(request, 'Nessun file caricato')
            return redirect('mainapp:import_listino')
        
        excel_file = request.FILES['file']
        print(excel_file)

        try:
            # Leggo il file in memoria
            file_in_memory = io.BytesIO(excel_file.read())
            
            # carico il workbook
            wb = load_workbook(filename=file_in_memory, read_only=True, data_only=True)
            
            # primo foglio
            sheet = wb.active
            
            # converto il foglio in un DataFrame pandas
            data = sheet.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)

            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip().str.upper()
            df['DATA'] = pd.to_datetime(df['DATA']).dt.strftime('%Y-%m-%d')

        except Exception as e:
            print("Errore nel leggere il file")
            print(f"Errore specifico: {type(e).__name__}, {str(e)}")
            messages.error(request, f'Errore nel leggere il file Excel: {str(e)}')
            return redirect('mainapp:import_listino')
        
        # Verifica che il file abbia tutte le colonne
        required_columns = ['FORNITORE', 'MAGAZZINO', 'MEZZO', 'TIPOLOGIA', 'PARTENZA', 'ARRIVO', 'COSTO', 'DATA', 'CONTOCONTABILE', 'VOCESPESA', 'CENTROCOSTO']
        if not all(col in df.columns for col in required_columns):
            messages.error(request, 'Il file non contiene tutte le colonne richieste')
            return redirect('mainapp:import_listino')
        
        righe_inserite = 0
        righe_aggiornate = 0
        righe_fallite = 0
        righe_uguali = 0
    
        for _, row in df.iterrows():
            try:
                with transaction.atomic():
                    # Cerco di ottenere le entità esistenti
                    try:
                        fornitore = Fornitore.objects.get(fornitore_nome=row['FORNITORE'])
                        magazzino = Magazzino.objects.get(magazzino_lettera=row['MAGAZZINO'])
                        mezzo = Mezzo.objects.get(mezzo_nome=row['MEZZO'])
                        tipologia = Tipologia.objects.get(tipologia_nome=row['TIPOLOGIA'])
                        partenza = Zona.objects.get(zona_nome=row['PARTENZA'])
                        arrivo = Zona.objects.get(zona_nome=row['ARRIVO'])
                    except ObjectDoesNotExist as e:
                        errore_dettagliato = f"Entità non trovata: {type(e).__name__} - {str(e)}. "
                        errore_dettagliato += f"Valori riga: FORNITORE={row['FORNITORE']}, "
                        errore_dettagliato += f"MAGAZZINO={row['MAGAZZINO']}, MEZZO={row['MEZZO']}, "
                        errore_dettagliato += f"TIPOLOGIA={row['TIPOLOGIA']}, PARTENZA={row['PARTENZA']}, "
                        errore_dettagliato += f"ARRIVO={row['ARRIVO']}"
                        raise ValueError(errore_dettagliato)

                    data_della_riga = row['DATA']
                    if isinstance(data_della_riga, str):
                        data_della_riga = date.fromisoformat(data_della_riga)
                    elif isinstance(data_della_riga, (datetime, pd.Timestamp)):
                        data_della_riga = data_della_riga.date()
                    else:
                        raise ValueError(f"Formato data non valido: {data_della_riga}")

                    try:
                        riga_listino = Listino.objects.get(
                            fornitore=fornitore,
                            magazzino=magazzino,
                            mezzo=mezzo,
                            tipologia=tipologia,
                            partenza=partenza,
                            arrivo=arrivo
                        )

                        if riga_listino.data_ultimo_aggiornamento == data_della_riga:
                            print(f'Oggetto id={riga_listino.id} ha la data uguale. data oggetto = {riga_listino.data_ultimo_aggiornamento}, data riga ={data_della_riga} ')
                            righe_uguali += 1 
                            continue
                        else:
                            print(f"due date diverse. Data oggetto {riga_listino.data_ultimo_aggiornamento}, data riga{data_della_riga} ")
                            riga_listino.costo = row['COSTO']
                            riga_listino.data_ultimo_aggiornamento = data_della_riga
                            riga_listino.conto_contabile = row['CONTOCONTABILE']
                            riga_listino.voce_spesa =  row['VOCESPESA']
                            riga_listino.centro_costo = row['CENTROCOSTO']
                            riga_listino.save()
                            righe_aggiornate += 1
                    except Listino.DoesNotExist:
                        Listino.objects.create(
                            fornitore=fornitore,
                            magazzino=magazzino,
                            mezzo=mezzo,
                            tipologia=tipologia,
                            partenza=partenza,
                            arrivo=arrivo,
                            data_ultimo_aggiornamento=data_della_riga,
                            costo=row['COSTO'],
                            conto_contabile=row['CONTOCONTABILE'],
                            voce_spesa=row['VOCESPESA'],
                            centro_costo=row['CENTROCOSTO'])
                        righe_inserite += 1 
            except Exception as e:
                righe_fallite += 1
                dati_riga = row.to_dict()

                # Converto esplicitamente la colonna 'DATA' in stringa
                if 'DATA' in dati_riga:
                    if isinstance(dati_riga['DATA'], (pd.Timestamp, datetime.date)):
                        dati_riga['DATA'] = dati_riga['DATA'].strftime('%Y-%m-%d')
                    elif isinstance(dati_riga['DATA'], str):
                        # Già una stringa, non fare nulla
                        pass
                    else:
                        dati_riga['DATA'] = str(dati_riga['DATA'])  # Converti in stringa come fallback
                
                try:
                    with transaction.atomic():
                        InserimentoFallito.objects.create(
                            dati_riga=json.dumps(dati_riga, cls=CustomJSONEncoder),
                            errore=str(e)
                        )
                    print("Inserimento fallito creato")
                except Exception as inner_e:
                    print(f"Errore durante la creazione di InserimentoFallito: {str(inner_e)}")

        if righe_fallite > 0:
             messages.warning(request, f'Importazione completata. Inserite: {righe_inserite}, '
                                  f'Aggiornate: {righe_aggiornate}, Fallite: {righe_fallite}, Gia presenti: {righe_uguali}')
        else:    
            messages.success(request, f'Importazione completata. Inserite: {righe_inserite}, '
                                  f'Aggiornate: {righe_aggiornate}, Fallite: {righe_fallite}, Gia presenti: {righe_uguali}')
        return redirect('mainapp:import_listino')
    

class ListinoListView(ListView):
    model = Listino
    paginate_by = 40
    template_name = "mainapp/listini.html"
    context_object_name = "listini"

    def get_queryset(self):
        queryset = super().get_queryset()
        filters = {}
        
        filter_mappings = {
            'fornitore': 'fornitore__fornitore_nome',
            'magazzino': 'magazzino__magazzino_lettera',
            'mezzo': 'mezzo__mezzo_nome__icontains',
            'tipo': 'tipologia__tipologia_nome',
            'partenza': 'partenza__zona_nome__icontains',
            'arrivo': 'arrivo__zona_nome__icontains',
            'conto_contabile': 'conto_contabile__icontains',
            'voce_spesa': 'voce_spesa__icontains',
            'centro_costo': 'centro_costo__icontains',
        }

        for param, filter_name in filter_mappings.items():
            value = self.request.GET.get(param)
            if value:
                filters[filter_name] = value

        return queryset.filter(**filters).order_by('fornitore', 'magazzino')


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aggiungi i valori selezionati al contesto
        filter_fields = [
            'fornitore', 'magazzino', 'mezzo', 'tipo', 'partenza', 
            'arrivo', 'conto_contabile', 'voce_spesa', 'centro_costo'
        ]
        for field in filter_fields:
            context[field] = self.request.GET.get(field, '')

        # Aggiungi le opzioni per i dropdown
        context['fornitori'] = Fornitore.objects.values_list('fornitore_nome', flat=True).distinct()
        context['magazzini'] = Magazzino.objects.values_list('magazzino_lettera', flat=True).distinct()
        context['mezzi'] = Mezzo.objects.values_list('mezzo_nome', flat=True).distinct()
        context['tipi'] = Tipologia.objects.values_list('tipologia_nome', flat=True).distinct()
        context['partenze'] = Zona.objects.values_list('zona_nome', flat=True).distinct()
        context['arrivi'] = Zona.objects.values_list('zona_nome', flat=True).distinct()
        context['conti_contabili'] = Listino.objects.values_list('conto_contabile', flat=True).distinct()
        context['voci_spesa'] = Listino.objects.values_list('voce_spesa', flat=True).distinct()
        context['centri_costo'] = Listino.objects.values_list('centro_costo', flat=True).distinct()

        return context

# -------------------------------------------------------------------------------------------------------------- #
# ---------------------------- Gestione righe fallite  -------------------------- #
# -------------------------------------------------------------------------------------------------------------- #
class InserimentoFallitoListView(ListView):
    model = InserimentoFallito
    paginate_by = 20
    template_name = "mainapp/listino_falliti.html"
    context_object_name = "lista_falliti"
    

    def get_queryset(self):
        queryset = super().get_queryset()
        
        data_tentativo = self.request.GET.get('data_tentativo')
        tipo_errore = self.request.GET.get('tipo_errore')

        if data_tentativo:
            queryset = queryset.filter(data_tentativo__gte=data_tentativo)
        if tipo_errore:
            queryset = queryset.filter(errore__icontains=tipo_errore)

        return queryset.order_by('-data_tentativo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['data_tentativo'] = self.request.GET.get('data_tentativo', '')
        context['tipo_errore'] = self.request.GET.get('tipo_errore', '')
        return context

class EliminaInserimentoFallitoView(View):
    def post(self, request, pk):
        print("sono in questa post")
        print("PK = ", pk)
        oggetto = get_object_or_404(InserimentoFallito, pk=pk)
        oggetto.delete()
        messages.success(request, "L'inserimento fallito è stato eliminato con successo.")
        return redirect(reverse('mainapp:listino_falliti'))

class DeleteAllInserimentiFallitiView(View):
    def get(self, request):
        InserimentoFallito.objects.all().delete()
        messages.success(request, "Tutti gli inserimenti falliti sono stati cancellati con successo.")
        return redirect(reverse('mainapp:listino_falliti'))



def situazione_prefatture(request):
    # Inizializza il queryset con tutte le prefatture
    prefatture = Prefattura.objects.all().order_by('-fattura_data')
    
    # Applica i filtri
    numero_fattura = request.GET.get('numero_fattura')
    fornitore_id = request.GET.get('fornitore')
    data = request.GET.get('data')
    
    if numero_fattura:
        prefatture = prefatture.filter(fattura_numero=numero_fattura)
    if fornitore_id:
        prefatture = prefatture.filter(fattura_fornitore_id=fornitore_id)
    if data:
        prefatture = prefatture.filter(fattura_data=data)
    
    # Paginazione
    paginator = Paginator(prefatture, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    fornitori_list = Fornitore.objects.all().order_by('fornitore_nome')
    
    context = {
        'page_obj': page_obj,
        'fornitori_list': fornitori_list,
        'numero_fattura': numero_fattura,
        'fornitore_id': fornitore_id,
        'data': data,
    }
    
    return render(request, 'mainapp/situazione_prefatture.html', context)

def aggiungi_prefattura(request):
    if request.method == 'POST':
        numero = request.POST.get('fattura_numero')
        fornitore_id = request.POST.get('fattura_fornitore')
        data = request.POST.get('fattura_data')

        if numero == "" or numero is None:
            messages.error(request, "Errore: Il numero fattura è obbligatorio")
            return redirect('mainapp:situazione_prefatture')
        if data == "" or data is None:
            messages.error(request, "Errore: La data fattura è obbligatoria")
            return redirect('mainapp:situazione_prefatture')
        try:
            fornitore = Fornitore.objects.get(pk=fornitore_id)
            Prefattura.objects.create(
                fattura_numero=numero,
                fattura_fornitore=fornitore,
                fattura_data=data,
                fattura_utente=request.user
            )
            messages.success(request, 'Prefattura aggiunta con successo.')
        except Exception as e:
            messages.error(request, f'Errore nell\'aggiunta della prefattura: {str(e)}')

    return redirect('mainapp:situazione_prefatture')

def elimina_prefattura(request, pk):
    pass


def filter_righe_listino(request, pk):
    prefattura = get_object_or_404(Prefattura, pk=pk)
    righe_listino = Listino.objects.filter(fornitore=prefattura.fattura_fornitore)

    magazzino = request.GET.get('magazzino')
    mezzo = request.GET.get('mezzo')
    tipologia = request.GET.get('tipologia')
    partenza = request.GET.get('partenza')
    arrivo = request.GET.get('arrivo')

    if magazzino:
        righe_listino = righe_listino.filter(magazzino__magazzino_nome=magazzino)
    if mezzo:
        righe_listino = righe_listino.filter(mezzo__mezzo_nome=mezzo)
    if tipologia:
        righe_listino = righe_listino.filter(tipologia__tipologia_nome=tipologia)
    if partenza:
        righe_listino = righe_listino.filter(partenza__zona_nome=partenza)
    if arrivo:
        righe_listino = righe_listino.filter(arrivo__zona_nome=arrivo)

    data = []
    for riga in righe_listino:
        data.append({
            'id': riga.id,
            'magazzino': riga.magazzino.magazzino_lettera,
            'mezzo': riga.mezzo.mezzo_nome,
            'tipologia': riga.tipologia.tipologia_nome,
            'partenza': riga.partenza.zona_nome,
            'arrivo': riga.arrivo.zona_nome,
            'costo': riga.costo,
        })

    return JsonResponse(data, safe=False)

def edit_prefattura(request, pk):
    prefattura = get_object_or_404(Prefattura, pk=pk)
    righe_prefattura = PrefatturaRighe.objects.filter(prefattura=prefattura)

    magazzini = Magazzino.objects.values_list('id', 'magazzino_nome')
    mezzi = Mezzo.objects.values_list('id', 'mezzo_nome')
    tipologie = Tipologia.objects.values_list('id', 'tipologia_nome')
    partenze = Zona.objects.values_list('id', 'zona_nome')
    arrivi = Zona.objects.values_list('id', 'zona_nome')

    context = {
        'prefattura': prefattura,
        'righe_prefattura': righe_prefattura,
        'magazzini': magazzini,
        'mezzi': mezzi,
        'tipologie': tipologie,
        'partenze': partenze,
        'arrivi': arrivi,
    }

    if request.method == 'POST':
        listino_id = request.POST.get('listino_id')
        quantita = request.POST.get('quantita')

        if listino_id and quantita:
            listino = get_object_or_404(Listino, id=listino_id)
            PrefatturaRighe.objects.create(prefattura=prefattura, riga_listino=listino, quantita=quantita)
            return redirect('mainapp:edit_prefattura', pk=pk)

    return render(request, "mainapp/edit_prefattura.html", context)

def elimina_riga_prefattura(request, pk):
    pass

